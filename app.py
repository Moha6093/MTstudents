from flask import Flask, render_template, request
import openpyxl

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST', 'PUT'])
def index():
    if request.method == 'POST':
        # Get the form data
        student = request.form['student']
        phone_num1 = request.form['phone_num1']
        phone_num2 = request.form['phone_num2']
        grade = request.form['Grade']

        # Open the existing Excel file
        workbook = openpyxl.load_workbook('student_data.xlsx')
        sheet = workbook.active

        # Add the new data to the Excel file
        row = sheet.max_row + 1
        sheet.cell(row, 1).value = student
        sheet.cell(row, 2).value = phone_num1
        sheet.cell(row, 3).value = phone_num2
        sheet.cell(row, 4).value = grade

        # Save the Excel file
        workbook.save('student_data.xlsx')

        # Display the student data in a table
        data = [
            ['Name', 'Student Num', 'Parent Num', 'Grade'],
            [student, phone_num1, phone_num2, grade]
        ]
        return render_template('result.html', data=data)

    return render_template('index.html')

if __name__ == '__main__':
   app.run(debug=True, host='moha6093.github.io')
